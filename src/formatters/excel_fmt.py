# src/formatters/excel_fmt.py
from pathlib import Path
from typing import Sequence

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

from src.models import ArbitrationData, AttendanceRecord, AttendanceStats, MessageRecord

_RED_FONT = Font(color="FF0000", bold=True)
_YELLOW_FILL = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
_HEADER_FONT = Font(bold=True)


class ExcelFormatter:
    def __init__(self, output_dir: str = "./output"):
        self.output_dir = Path(output_dir)
        self.output_dir.mkdir(parents=True, exist_ok=True)

    def format(self, data: ArbitrationData) -> str:
        filename = f"仲裁证据_{data.username}_{data.period_start}-{data.period_end}.xlsx"
        filepath = self.output_dir / filename

        wb = Workbook()
        self._write_attendance_sheet(wb, data.attendance_records)
        self._write_stats_sheet(wb, data.attendance_stats)
        self._write_messages_sheet(wb, data.message_records)

        wb.save(filepath)
        return str(filepath)

    def _write_attendance_sheet(
        self, wb: Workbook, records: Sequence[AttendanceRecord]
    ) -> None:
        ws = wb.active
        ws.title = "考勤流水"

        headers = ["用户ID", "打卡时间", "打卡位置", "打卡类型", "打卡结果", "备注"]
        for col, h in enumerate(headers, 1):
            cell = ws.cell(1, col, h)
            cell.font = _HEADER_FONT

        anomaly_keywords = {"迟到", "早退", "缺卡", "缺勤", "异常"}
        for row, rec in enumerate(records, 2):
            ws.cell(row, 1, rec.user_id)
            ws.cell(row, 2, rec.check_time.strftime("%Y-%m-%d %H:%M:%S"))
            ws.cell(row, 3, rec.location_name)
            ws.cell(row, 4, rec.type)
            result_cell = ws.cell(row, 5, rec.check_result)
            ws.cell(row, 6, rec.comment)
            if rec.check_result in anomaly_keywords:
                result_cell.font = _RED_FONT
                for c in range(1, 7):
                    ws.cell(row, c).fill = _YELLOW_FILL

        summary_row = len(records) + 2
        ws.cell(summary_row, 1, "合计")
        ws.cell(summary_row, 2, f"共 {len(records)} 条记录")
        ws.cell(summary_row, 1).font = _HEADER_FONT

    def _write_stats_sheet(
        self, wb: Workbook, stats_list: Sequence[AttendanceStats]
    ) -> None:
        ws = wb.create_sheet("考勤统计")

        headers = [
            "用户ID", "姓名", "统计开始", "统计结束",
            "出勤天数", "迟到次数", "早退次数", "加班时长(h)", "请假天数",
        ]
        for col, h in enumerate(headers, 1):
            cell = ws.cell(1, col, h)
            cell.font = _HEADER_FONT

        total_days = 0
        total_late = 0
        total_overtime = 0.0
        for row, s in enumerate(stats_list, 2):
            ws.cell(row, 1, s.user_id)
            ws.cell(row, 2, s.user_name)
            ws.cell(row, 3, s.period_start)
            ws.cell(row, 4, s.period_end)
            ws.cell(row, 5, s.attendance_days)
            ws.cell(row, 6, s.late_count)
            ws.cell(row, 7, s.early_leave_count)
            ws.cell(row, 8, s.overtime_hours)
            ws.cell(row, 9, s.leave_days)
            total_days += s.attendance_days
            total_late += s.late_count
            total_overtime += s.overtime_hours

        summary_row = len(stats_list) + 2
        ws.cell(summary_row, 1, "合计")
        ws.cell(summary_row, 5, total_days)
        ws.cell(summary_row, 6, total_late)
        ws.cell(summary_row, 8, total_overtime)
        ws.cell(summary_row, 1).font = _HEADER_FONT

    def _write_messages_sheet(
        self, wb: Workbook, messages: Sequence[MessageRecord]
    ) -> None:
        ws = wb.create_sheet("沟通记录")

        headers = ["聊天ID", "发送者ID", "消息类型", "发送时间", "消息内容"]
        for col, h in enumerate(headers, 1):
            cell = ws.cell(1, col, h)
            cell.font = _HEADER_FONT

        for row, msg in enumerate(messages, 2):
            ws.cell(row, 1, msg.chat_id)
            ws.cell(row, 2, msg.sender_id)
            ws.cell(row, 3, msg.msg_type)
            ws.cell(row, 4, msg.create_time.strftime("%Y-%m-%d %H:%M:%S"))
            ws.cell(row, 5, msg.content)

        summary_row = len(messages) + 2
        ws.cell(summary_row, 1, "合计")
        ws.cell(summary_row, 2, f"共 {len(messages)} 条消息")
        ws.cell(summary_row, 1).font = _HEADER_FONT
