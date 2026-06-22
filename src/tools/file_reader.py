import csv
import io
from pathlib import Path

import openpyxl


def _read_excel_as_text(file_path: str, max_lines: int = 500) -> str:
    """Read Excel file and convert to readable text for AI analysis."""
    wb = openpyxl.load_workbook(file_path, data_only=True)
    sections = []
    line_count = 0

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        lines = [f"=== Sheet: {sheet_name} ==="]

        for row in ws.iter_rows(values_only=True):
            if line_count >= max_lines:
                lines.append("... [Data truncated due to size limit. Please summarize or chunk further] ...")
                break

            # Skip completely empty rows
            if all(v is None for v in row):
                continue

            # Format each cell
            formatted_cells = []
            for cell in row:
                if cell is None:
                    formatted_cells.append("")
                elif isinstance(cell, (int, float)):
                    formatted_cells.append(str(cell))
                elif hasattr(cell, 'isoformat'):  # datetime
                    formatted_cells.append(cell.strftime("%Y-%m-%d %H:%M"))
                else:
                    formatted_cells.append(str(cell).strip())

            lines.append(" | ".join(formatted_cells))
            line_count += 1
            
        sections.append("\n".join(lines))
        if line_count >= max_lines:
            break

    return "\n\n".join(sections)


def _read_csv_as_text(file_path: str, max_lines: int = 500) -> str:
    """Read CSV file and convert to readable text for AI analysis."""
    content = Path(file_path).read_text(encoding="utf-8")
    lines = []
    line_count = 0

    reader = csv.reader(io.StringIO(content))
    for row in reader:
        if line_count >= max_lines:
            lines.append("... [Data truncated due to size limit. Please summarize or chunk further] ...")
            break
        lines.append(" | ".join(str(cell).strip() for cell in row))
        line_count += 1

    return "\n".join(lines)


def parse_attendance_file(file_path: str, user_hint: str | None = None, max_lines: int = 500) -> str:
    """Read attendance file and return as structured text for AI analysis.

    This function does NOT parse or interpret the data — it only converts
    the raw file content into clean, structured text that a Claude Code Agent
    can directly analyze using its built-in AI capabilities.

    Args:
        file_path: Path to the attendance file (CSV or Excel).
        user_hint: Optional instruction to guide AI analysis, e.g.
            "重点关注加班费和迟到记录" or "检查是否有缺卡情况"
        max_lines: Maximum number of rows to parse to prevent prompt limits.

    Returns:
        Structured text representation of the file content,
        suitable for AI analysis. The Agent should interpret this
        and extract: AttendanceRecord, AttendanceStats, anomalies, etc.
    """
    path = Path(file_path)

    if path.suffix.lower() == ".csv":
        raw_text = _read_csv_as_text(file_path, max_lines)
    elif path.suffix.lower() in (".xlsx", ".xls"):
        raw_text = _read_excel_as_text(file_path, max_lines)
    else:
        raise ValueError(f"Unsupported file format: {path.suffix}. Use .csv or .xlsx")

    # Add header context
    header = [
        "=" * 60,
        "[Attendance File Parsed - AI Analysis Ready]",
        f"File: {path.name}",
        f"Format: {path.suffix.upper()}",
        "=" * 60,
        "",
    ]

    if user_hint:
        header.extend([
            f"[User Focus: {user_hint}]",
            "",
        ])

    header.append("--- Raw Data Start ---")
    footer = ["--- Raw Data End ---"]

    return "\n".join(header) + "\n" + raw_text + "\n" + "\n".join(footer)
