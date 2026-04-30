from pathlib import Path

from src.models import AttendanceRecord, AttendanceStats
from .csv_parser import _parse_daily_flow, _parse_monthly_stats


def parse_attendance_excel(file_path: str) -> tuple[list[AttendanceRecord], list[AttendanceStats]]:
    """
    Parse attendance Excel file.
    Supports both daily flow (sheet: 打卡流水) and monthly stats (sheet: 考勤统计).
    Returns (records, stats).
    """
    import openpyxl

    wb = openpyxl.load_workbook(file_path, data_only=True)
    records = []
    stats = []

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))

        if not rows:
            continue

        # Get headers from first row
        headers = [str(h).strip() if h is not None else "" for h in rows[0]]
        data_rows = [dict(zip(headers, row)) for row in rows[1:] if any(v is not None for v in row)]

        # Detect format by headers
        header_set = set(headers)
        if "日期" in header_set or "时间" in header_set:
            records.extend(_parse_daily_flow(data_rows))
        elif "出勤天数" in header_set or "late_count" in header_set or "attendance_days" in header_set:
            stats.extend(_parse_monthly_stats(data_rows))
        else:
            # Try both: first data row decides
            first = data_rows[0] if data_rows else {}
            if any("日期" in str(k) or "时间" in str(k) for k in first.keys()):
                records.extend(_parse_daily_flow(data_rows))
            else:
                stats.extend(_parse_monthly_stats(data_rows))

    return records, stats
